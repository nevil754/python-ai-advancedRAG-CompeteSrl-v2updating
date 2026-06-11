# =============================================================
# app/rag/ingestion/parser.py
# Parsing documenti: docling per PDF/DOCX complessi,
# unstructured come fallback, pypdf per PDF semplici.
# =============================================================

from __future__ import annotations  #x python legacy in prj big soprattutto, trasforma 'def get_user()->User:' in 'def get_user() -> "User":' quindi tutte le annotazioni vengono conservate come str
import mimetypes   #x indovinare il MIME type di un file, e.g. file .html -> "text/html", file .jpg -> "image/jpeg"
from pathlib import Path
from typing import Any
from loguru import logger
from app.core.settings import get_settings


settings = get_settings()

class ParsedDocument:
    """Risultato del parsing: testo strutturato + metadata estratti."""
    def __init__(
        self,
        text: str,
        pages: list[str],        #testo diviso per pagina
        tables: list[dict],      #tabelle estratte come dict
        metadata: dict[str, Any],
        page_count: int,
    ):
        self.text = text
        self.pages = pages
        self.tables = tables
        self.metadata = metadata
        self.page_count = page_count


def parse_document(file_path: str) -> ParsedDocument:
    """
    Parsa un documento scegliendo il parser migliore per il tipo di file.
    Strategia:
    - PDF/DOCX → docling (preserva struttura, tabelle, intestazioni)
    - Fallback → unstructured (più robusto ma meno preciso)
    - XLSX → openpyxl diretto
    - TXT/MD → lettura diretta
    Args:
        file_path: path assoluto del file
    Returns:
        ParsedDocument con testo e metadata estratti
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File non trovato: {file_path}")
    suffix = path.suffix.lower()  #.suffix è property di pathlib.Path che return l'estensione del file e.g. pdf / docx / ect
    mime_type = mimetypes.guess_type(file_path)[0] or ""  # guess_type() ritorna una tupla (type, encoding), con [0] prendo solo il type
    logger.info(f"Parsing documento: {path.name} ({suffix})")
    if suffix in {".pdf", ".docx", ".pptx"} and settings.ingestion_prefer_docling:
        try:
            return _parse_with_docling(file_path)
        except Exception as e:
            logger.warning(f"Docling fallito ({e}), fallback su unstructured")
            return _parse_with_unstructured(file_path)
    elif suffix in {".xlsx", ".xls"}:
        return _parse_excel(file_path)
    elif suffix in {".txt", ".md"}:
        return _parse_text(file_path)
    else:
        return _parse_with_unstructured(file_path)

def _parse_with_docling(file_path: str) -> ParsedDocument:
    """
    Parser avanzato con docling.
    Preserva struttura documenti complessi: contratti, bilanci, normative.
    Estrae tabelle come testo strutturato markdown.
    """
    from docling.document_converter import DocumentConverter
    from docling.datamodel.base_models import InputFormat
    from docling.datamodel.pipeline_options import PdfPipelineOptions  #obj di config per la pipeline pdf 
    pipeline_options = PdfPipelineOptions()
    pipeline_options.do_table_structure = settings.ingestion_extract_tables  #do_table_structure è booleano: True  → estrai tabelle strutturate, False → non farlo (o estrazione semplificata)
    pipeline_options.do_ocr = False  #False disattivo il riconoscimento del testo dalle immagini (OCR, xk è leento)
    converter = DocumentConverter()
    result = converter.convert(file_path)   #🔥lo converte in un oggetto docling Document, con testo, pagine, tabelle e metadata estratti
    doc = result.document
    # Esporta in markdown — preserva intestazioni, tabelle, liste
    full_text = doc.export_to_markdown()  #ora hai ottenuto la versione Markdown!!
    pages = []
    tables = []
    for page_num, page in enumerate(doc.pages, 1):
        page_text = ""
        for element in page.get_elements():
            page_text += element.text + "\n"
        pages.append(page_text)
    if settings.ingestion_extract_tables:  #questo field esiste in config.yaml
        for table in doc.tables:
            tables.append({
                "page": table.prov[0].page_no if table.prov else None,
                "markdown": table.export_to_markdown(),
            })
    metadata = {
        "parser": "docling",
        "page_count": len(doc.pages),
        "table_count": len(tables),
        "has_tables": len(tables) > 0,
    }
    logger.debug(
        "Docling parsing completato",
        pages=len(pages),
        tables=len(tables),
        chars=len(full_text),
    )
    return ParsedDocument(
        text=full_text,
        pages=pages,
        tables=tables,
        metadata=metadata,
        page_count=len(pages),
    )

def _parse_with_unstructured(file_path: str) -> ParsedDocument:
    """
    Parser generale con unstructured. Fallback per tutti i formati.
    Meno preciso di docling ma supporta più tipi di file.
    """
    from unstructured.partition.auto import partition
    elements = partition(
        filename=file_path,
        include_page_breaks=True,
        strategy="fast",
    )
    pages: list[str] = []
    current_page: list[str] = []
    tables: list[dict] = []
    for el in elements:
        el_type = type(el).__name__
        text = str(el)
        if el_type == "PageBreak":
            pages.append("\n".join(current_page))
            current_page = []
        elif el_type == "Table":
            tables.append({"text": text, "page": len(pages) + 1})
            current_page.append(text)
        else:
            current_page.append(text)
    if current_page:
        pages.append("\n".join(current_page))
    full_text = "\n\n".join(pages)
    return ParsedDocument(
        text=full_text,
        pages=pages,
        tables=tables,
        metadata={"parser": "unstructured", "page_count": len(pages)},
        page_count=len(pages),
    )

def _parse_excel(file_path: str) -> ParsedDocument:
    """Estrae testo da file Excel come tabelle markdown."""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets_text: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append(" | ".join(str(c or "") for c in row))
        if rows:
            sheet_md = f"## Foglio: {sheet_name}\n\n" + "\n".join(rows)
            sheets_text.append(sheet_md)
    full_text = "\n\n".join(sheets_text)
    return ParsedDocument(
        text=full_text,
        pages=[full_text],
        tables=[],
        metadata={"parser": "openpyxl", "sheets": wb.sheetnames},
        page_count=1,
    )

def _parse_text(file_path: str) -> ParsedDocument:
    """Lettura diretta per file .txt e .md."""
    with open(file_path, encoding="utf-8", errors="replace") as f:
        text = f.read()
    return ParsedDocument(
        text=text,
        pages=[text],
        tables=[],
        metadata={"parser": "text"},
        page_count=1,
    )

